# audit/exports.py
import io
from datetime import datetime

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.mixins import ModulePermissionMixin
from core.pdf_utils import render_pdf_response
from .models import AuditPlan, AuditExecution, AuditSchedule


def _thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _hs():
    return {"font": Font(bold=True, color="FFFFFF", size=10),
            "fill": PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid"),
            "alignment": Alignment(vertical="center", wrap_text=True), "border": _thin_border()}


def _cs():
    return {"font": Font(size=10), "alignment": Alignment(vertical="center", wrap_text=True), "border": _thin_border()}


def _apply(cell, styles):
    for attr, val in styles.items():
        setattr(cell, attr, val)


# ─── AuditPlan PDF + Excel ────────────────────────────────────────────────────

class AuditPlanPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "audit.view_auditplan"

    def get(self, request, pk):
        plan = get_object_or_404(
            AuditPlan.objects.select_related("approved_by").prefetch_related("schedules__auditor"),
            pk=pk,
        )
        return render_pdf_response(
            "audit/pdf_plan.html",
            {"plan": plan, "now": datetime.now()},
            f"plan_auditoria_{plan.year}.pdf",
            user=request.user,
            source_model="audit.AuditPlan",
            source_id=plan.pk,
            source_code=f"AUDIT-PLAN-{plan.year}",
            data_snapshot={
                "name": plan.name,
                "year": plan.year,
                "status": plan.status,
                "schedules": plan.schedules.count(),
            },
        )


class AuditPlanExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "audit.view_auditplan"

    def get(self, request, pk):
        plan = get_object_or_404(AuditPlan.objects.select_related("approved_by").prefetch_related("schedules__auditor"), pk=pk)
        wb = Workbook(); ws = wb.active; ws.title = f"Plan {plan.year}"
        hs = _hs(); cs = _cs()
        ws.merge_cells("A1:E1")
        ws["A1"] = f"Plan de Auditoria — {plan.name} ({plan.year})"
        ws["A1"].font = Font(bold=True, size=13, color="1F2937")
        info = [
            ("Nombre", plan.name), ("Ano", plan.year), ("Estado", plan.get_status_display()),
            ("Aprobado por", plan.approved_by.get_full_name() if plan.approved_by else "—"),
            ("Alcance", plan.scope or "—"),
        ]
        r = 3
        for lbl, val in info:
            c = ws.cell(row=r, column=1, value=lbl); _apply(c, hs)
            v = ws.cell(row=r, column=2, value=str(val)); _apply(v, cs); r += 1
        r += 1
        for col, h in enumerate(["Proceso", "Auditor", "Fecha planificada", "Responsable area", "Estado"], 1):
            cell = ws.cell(row=r, column=col, value=h); _apply(cell, hs)
        r += 1
        for s in plan.schedules.all():
            for col, v in enumerate([s.get_process_display(), s.auditor.get_full_name() or s.auditor.username, str(s.planned_date), s.area_responsible, s.get_status_display()], 1):
                cell = ws.cell(row=r, column=col, value=v); _apply(cell, cs)
            r += 1
        for i, w in enumerate([25, 22, 16, 22, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="plan_auditoria_{plan.year}.xlsx"'
        return resp


# ─── AuditExecution PDF + Excel ──────────────────────────────────────────────

class AuditExecutionPDFView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "audit.view_auditexecution"

    def get(self, request, pk):
        execution = get_object_or_404(
            AuditExecution.objects.select_related(
                "schedule__plan", "schedule__auditor", "auditor", "auditee",
            ).prefetch_related("findings__nc"),
            pk=pk,
        )
        try:
            closure = execution.closure
        except Exception:
            closure = None
        return render_pdf_response(
            "audit/pdf_execution.html",
            {"execution": execution, "closure": closure,
             "summary": execution.finding_summary, "now": datetime.now()},
            f"auditoria_{execution.schedule.plan.year}_{execution.pk}.pdf",
            user=request.user,
            source_model="audit.AuditExecution",
            source_id=execution.pk,
            source_code=f"AUDIT-{execution.schedule.plan.year}-{execution.pk}",
            data_snapshot={
                "process": execution.schedule.process,
                "auditor": execution.auditor.username if execution.auditor else "",
                "actual_date": str(execution.actual_date),
                "findings": execution.findings.count(),
                "closed": bool(closure),
            },
        )


class AuditExecutionExcelView(LoginRequiredMixin, ModulePermissionMixin, View):
    permission_required = "audit.view_auditexecution"

    def get(self, request, pk):
        execution = get_object_or_404(AuditExecution.objects.select_related("schedule__plan", "auditor", "auditee").prefetch_related("findings"), pk=pk)
        wb = Workbook(); ws = wb.active; ws.title = "Ejecucion Auditoria"
        hs = _hs(); cs = _cs()
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Ejecucion de Auditoria — {execution.schedule.get_process_display()}"
        ws["A1"].font = Font(bold=True, size=13, color="1F2937")
        info = [
            ("Plan", execution.schedule.plan.name),
            ("Proceso auditado", execution.schedule.get_process_display()),
            ("Auditor", execution.auditor.get_full_name() or execution.auditor.username),
            ("Auditado", execution.auditee.get_full_name() or execution.auditee.username),
            ("Fecha ejecucion", str(execution.actual_date)),
        ]
        r = 3
        for lbl, val in info:
            c = ws.cell(row=r, column=1, value=lbl); _apply(c, hs)
            v = ws.cell(row=r, column=2, value=val); _apply(v, cs); r += 1
        r += 1
        for col, h in enumerate(["Tipo hallazgo", "Clausula ISO", "Descripcion", "NC vinculada"], 1):
            cell = ws.cell(row=r, column=col, value=h); _apply(cell, hs)
        r += 1
        for f in execution.findings.all():
            for col, v in enumerate([f.get_finding_type_display(), f.iso_clause or "—", f.description, f.nc.code if f.nc else "—"], 1):
                cell = ws.cell(row=r, column=col, value=v); _apply(cell, cs)
            r += 1
        for i, w in enumerate([18, 16, 50, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="auditoria_{execution.pk}.xlsx"'
        return resp
