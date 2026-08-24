"""
Armado de documentos Excel con el mismo formato en todos los modulos.

Los exports de documento (factura, despacho, orden de produccion, expediente
de lote...) comparten la misma estructura: titulo, bloques de datos con sus
etiquetas, tablas de detalle y, cuando aplica, un bloque de totales. Esta
clase concentra ese armado para que los reportes se vean igual entre modulos
y para que cada vista se limite a decir que datos van, no como se dibujan.

Uso:

    doc = ExcelDoc("FACTURA — 001-001-000000123")
    doc.seccion("EMISOR")
    doc.datos([("R.U.C.", cfg.ruc), ("Razon social", cfg.legal_name)])
    doc.tabla(["Producto", "Cantidad"], [[p.name, p.qty] for p in lineas],
              numericas=[2])
    doc.totales([("VALOR TOTAL", inv.total_amount)])
    return doc.response(f"factura_{inv.full_number}.xlsx")
"""
import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

AZUL = "134D5F"
GRIS_TITULO = "1F2937"
MONEDA = "#,##0.00"


def _borde():
    lado = Side(style="thin")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


class ExcelDoc:
    """Hoja de un documento, con secciones, bloques de datos y tablas."""

    def __init__(self, titulo, hoja="Documento", columnas=6):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = hoja[:31]
        self.columnas = columnas
        self.fila = 1
        self._titulo(titulo)

    # ── bloques ──────────────────────────────────────────────────────────────

    def _titulo(self, texto):
        ws = self.ws
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=self.columnas
        )
        c = ws.cell(row=1, column=1, value=texto)
        c.font = Font(bold=True, size=13, color=GRIS_TITULO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        self.fila = 3

    def seccion(self, titulo):
        """Franja de encabezado que separa un bloque del siguiente."""
        ws = self.ws
        ws.merge_cells(
            start_row=self.fila, start_column=1,
            end_row=self.fila, end_column=self.columnas,
        )
        c = ws.cell(row=self.fila, column=1, value=titulo)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
        c.alignment = Alignment(horizontal="left", vertical="center")
        self.fila += 1

    def datos(self, pares, espacio=True):
        """Pares etiqueta/valor, uno por fila. Los vacios salen como guion."""
        ws = self.ws
        for etiqueta, valor in pares:
            c = ws.cell(row=self.fila, column=1, value=etiqueta)
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill(start_color="EEF2F5", end_color="EEF2F5", fill_type="solid")
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = _borde()

            ws.merge_cells(
                start_row=self.fila, start_column=2,
                end_row=self.fila, end_column=self.columnas,
            )
            v = ws.cell(
                row=self.fila, column=2,
                value=valor if valor not in (None, "") else "—",
            )
            v.font = Font(size=10)
            v.alignment = Alignment(vertical="center", wrap_text=True)
            v.border = _borde()
            self.fila += 1
        if espacio:
            self.fila += 1

    def tabla(self, cabeceras, filas, numericas=(), vacia="Sin registros"):
        """Tabla de detalle. `numericas` son columnas 1-based con formato moneda."""
        ws = self.ws
        for col, texto in enumerate(cabeceras, 1):
            c = ws.cell(row=self.fila, column=col, value=texto)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _borde()
        self.fila += 1

        if not filas:
            ws.merge_cells(
                start_row=self.fila, start_column=1,
                end_row=self.fila, end_column=max(len(cabeceras), 1),
            )
            c = ws.cell(row=self.fila, column=1, value=vacia)
            c.font = Font(size=10, italic=True, color="6B7280")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _borde()
            self.fila += 2
            return

        for valores in filas:
            for col, valor in enumerate(valores, 1):
                c = ws.cell(
                    row=self.fila, column=col,
                    value=valor if valor not in (None, "") else "—",
                )
                c.font = Font(size=10)
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.border = _borde()
                if col in numericas:
                    c.number_format = MONEDA
                    c.alignment = Alignment(horizontal="right", vertical="center")
            self.fila += 1
        self.fila += 1

    def totales(self, pares, destacar=None):
        """Totales alineados a la derecha; `destacar` va en negrita."""
        ws = self.ws
        etiqueta_col = max(self.columnas - 1, 1)
        for etiqueta, valor in pares:
            c = ws.cell(row=self.fila, column=etiqueta_col, value=etiqueta)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill(start_color=AZUL, end_color=AZUL, fill_type="solid")
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _borde()

            v = ws.cell(row=self.fila, column=self.columnas, value=valor)
            v.number_format = MONEDA
            v.alignment = Alignment(horizontal="right", vertical="center")
            v.border = _borde()
            v.font = Font(bold=True, size=11) if etiqueta == destacar else Font(size=10)
            self.fila += 1

    def responsables(self, firmas):
        """Bloque de responsables (recibe/revisa/aprueba/libera...) al pie.

        `firmas` es la lista de core.responsibilities.build_signatories(obj).
        Si esta vacia no dibuja nada.
        """
        if not firmas:
            return
        self.seccion("RESPONSABLES")
        self.tabla(
            ["Rol", "Responsable", "Fecha"],
            [[f["role"], f["name"], f["date"]] for f in firmas],
            vacia="Sin responsables registrados",
        )

    def texto(self, contenido, etiqueta=None):
        """Parrafo largo (observaciones, conclusiones) en una fila combinada."""
        if etiqueta:
            self.datos([(etiqueta, contenido)], espacio=True)
            return
        ws = self.ws
        ws.merge_cells(
            start_row=self.fila, start_column=1,
            end_row=self.fila, end_column=self.columnas,
        )
        c = ws.cell(row=self.fila, column=1, value=contenido or "—")
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = _borde()
        self.fila += 2

    # ── cierre ───────────────────────────────────────────────────────────────

    def anchos(self, medidas):
        for i, ancho in enumerate(medidas, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = ancho

    def hoja(self, titulo, columnas=None):
        """Abre una hoja adicional dentro del mismo libro."""
        self.ws = self.wb.create_sheet(titulo[:31])
        if columnas:
            self.columnas = columnas
        self.fila = 1
        return self.ws

    def response(self, nombre_archivo, source=None):
        """Cierra el libro y devuelve la descarga.

        Si se pasa `source` (el objeto del documento), agrega al pie el bloque
        de responsables (recibe/revisa/aprueba/libera...) tomado de sus campos.
        """
        if source is not None:
            from core.responsibilities import build_signatories

            self.responsables(build_signatories(source))
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.getvalue(), content_type=XLSX_CONTENT_TYPE)
        resp["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
        return resp


def fecha(valor, formato="%d/%m/%Y"):
    """Formatea fechas/None de forma uniforme."""
    return valor.strftime(formato) if valor else "—"


def si_no(valor):
    return "SI" if valor else "NO"
